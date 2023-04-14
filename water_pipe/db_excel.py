from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, borders, colors
from pathlib import Path

from water_pipe.base import Connect, DTYPE

class ExcelConnect(Connect):
    def __init__(self, norm_config) -> None:
        """
        norm_config = {
            "path": "/home/",
            "filename": "data",
        }
        """
        config = norm_config.copy()
        
        path = config.get("path") if config.get("path") else "."
        filename = config["filename"]
        
        self.file = Path(path).joinpath(filename)
        
        ##############################
        self.connect = None
        self.cursor = None
        # wb.guess_types = True  # 猜测格式类型
        
        self.std_schema_data = []
        self.dataset_comment = ""
        self.placeholders = ""
        ##############################
    
    def _add_extend_method(self):
        """
        cursor  扩展 close 方法
        connect 扩展 commit 和 close 方法
        """
        # 给 self.cursor(Worksheet) 增加 close 方法
        from types import MethodType
        def close(self): pass
        self.cursor.close = MethodType(close, self.cursor)
        
        # 给 self.connect(Workbook) 增加 commit 方法
        # wb.save(file)
        def commit(self, file=self.file):
            self.save(file)
        self.connect.commit = MethodType(commit, self.connect)
        
        # Workbook 已经存在 close 方法，所以无需增加
        
    def execute(self, sql):
        raise Exception("Not Supported")
        
    def query(self, sql):
        self.set_query_schema()        
        # self.data = iter(row for row in self.cursor.values)
        self.data = iter(row for row in self.cursor.iter_rows(min_row=2, values_only=True))  # 跳过首行
        return self.fetchmany
    
    def set_query_schema(self):
        """ return dataset schema: col_name, data_type, comment"""
        self.connect = load_workbook(self.file)
        self.cursor = self.connect.active
        self._add_extend_method()
        
        cols_name = self.cursor.values.__next__()  # 获取第一行数据
        for col_name in cols_name:
            self.std_schema_data.append([col_name, DTYPE("text"), ""])
        self.placeholders = ",".join(["%s"] * len(self.std_schema_data))
    
    def fetchmany(self, size):
        batch_data = []
        for _ in range(size):
            try:
                batch_data.append(next(self.data))
            except StopIteration:
                break
        return batch_data

    def table(self, table_name):
        raise Exception("Not Supported")
    
    def insert(self, data):
        for row in data:
            self.cursor.append(row)
    
    def create_table(self, header=True): 
        self.connect = Workbook()
        self.cursor = self.connect.active
        self._add_extend_method()
        
        if header:
            header_data = [x[0] for x in self.std_schema_data]
            self.cursor.append(header_data)

            # font = Font(name='微软雅黑', size=9, bold=True)
            # border = Border(  # 边框
            #     top=Side(style="thin"),
            #     bottom=Side(style="thin"),
            #     left=Side(style="thin"),
            #     right=Side(style="thin")
            # )
            # for i in range(len(header_data)):
            #     self.cursor.cell(1, i + 1).font = font
            #     self.cursor.cell(1, i + 1).border = border
